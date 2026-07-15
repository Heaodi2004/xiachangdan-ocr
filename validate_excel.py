import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('c:/Users/he/Desktop/下场单/output/2026-07-15_北京测试公司.xlsx')
ws = wb['空白表']

print('=== 生成的Excel验证 ===')
print('Sheet names:', wb.sheetnames)
print('Dimensions:', ws.dimensions)

print('\n=== 表头信息 ===')
print('A1:', ws['A1'].value)
print('G1:', ws['G1'].value)
print('H1:', ws['H1'].value)
print('A2:', ws['A2'].value)
print('G2:', ws['G2'].value)
print('H2:', ws['H2'].value)

print('\n=== 颜色验证 ===')
for coord in ['I2', 'I3', 'I4']:
    cell = ws[coord]
    fill = cell.fill
    print(f'{coord} ({cell.value}): fgColor={fill.fgColor.rgb}')

print('\n=== 设备行颜色 ===')
for row in range(6, 9):
    cell = ws.cell(row=row, column=1)
    fill = cell.fill
    print(f'Row {row}: fgColor={fill.fgColor.rgb}')

print('\n=== 设备信息 ===')
for row in range(6, 9):
    print(f'Row {row}: 序号={ws.cell(row,1).value}, 设备={ws.cell(row,3).value}, 型号={ws.cell(row,4).value}, 编号={ws.cell(row,5).value}')

print('\n=== 汇总信息 ===')
print('设备总数:', ws['G4'].value)
print('带回数量:', ws['H4'].value)