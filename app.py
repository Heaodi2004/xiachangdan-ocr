import os
import re
import csv
import json
import cv2
import uuid
import time
import numpy as np
from flask import Flask, request, jsonify, render_template, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

os.environ['KMP_DUPLICATE_LIB_OK'] = 'TRUE'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'output')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp', 'tiff', 'tif'}
MAX_FILE_SIZE = 10 * 1024 * 1024
DEBUG_MODE = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
OCR_LAZY_INIT = os.environ.get('OCR_LAZY_INIT', 'True').lower() == 'true'
USE_LLM = os.environ.get('USE_LLM', 'True').lower() == 'true'
LLM_MODEL = os.environ.get('LLM_MODEL', 'qwen3.7-plus')

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

ocr_reader = None
ocr_engine = 'easyocr'
ocr_init_started = False

llm_config = {
    'api_key': '',
    'base_url': '',
    'model': LLM_MODEL,
    'enabled': False
}

def load_llm_config():
    global llm_config
    csv_path = os.path.join(BASE_DIR, 'api.csv')
    if not os.path.exists(csv_path):
        print('api.csv not found, LLM disabled')
        return
    
    try:
        config = {}
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) >= 2:
                    config[row[0].strip()] = row[1].strip()
        
        api_key = config.get('apiKey', '')
        base_url = config.get('openAiCompatible', '')
        
        if api_key and base_url:
            llm_config['api_key'] = api_key
            llm_config['base_url'] = base_url
            llm_config['enabled'] = True
            llm_config['model'] = LLM_MODEL
            print(f'LLM config loaded: model={LLM_MODEL}')
        else:
            print('LLM config incomplete, LLM disabled')
    except Exception as e:
        print(f'Failed to load LLM config: {e}')

def init_ocr():
    global ocr_reader, ocr_init_started
    if ocr_reader is None and not ocr_init_started:
        ocr_init_started = True
        try:
            import easyocr
            ocr_reader = easyocr.Reader(['ch_sim', 'en'], gpu=False, verbose=False)
            print('EasyOCR initialized successfully')
        except Exception as e:
            print(f'EasyOCR init failed: {e}')
            ocr_reader = None
            ocr_init_started = False
    return ocr_reader

if not OCR_LAZY_INIT:
    init_ocr()

load_llm_config()

def call_llm_extract(ocr_text, category='inspected'):
    if not USE_LLM or not llm_config['enabled']:
        return None, 0
    
    try:
        import requests
        
        category_name = {
            'inspected': '外检设备',
            'brought_back': '带回设备',
            'returned': '退检设备'
        }.get(category, '外检设备')
        
        system_prompt = """你是一个专业的设备信息提取助手。请从用户提供的OCR识别文本中，提取设备相关信息，并严格按照JSON格式返回。

需要提取的字段（对应Excel表格表头）：
1. cert_date - 证书日期（检定/校准日期，格式YYYY-MM-DD，如果没有则留空）
2. device_name - 设备名称（仪器名称、产品名称）
3. model - 规格型号（型号规格、型号、Type、Model）
4. factory_number - 出厂编号（出厂号、序列号、SN、Serial No.、编号）
5. device_number - 设备编号（资产编号、内部编号，如果没有则留空）
6. manufacturer - 生产厂家（制造厂家、制造商、品牌、Manufacturer）
7. remark - 设备备注（其他重要信息：精度、量程、校准点、使用地点、温湿度要求等，如果没有则留空）
8. responsible_person - 负责人（如果有则提取，没有则留空）

要求：
- 只返回JSON格式，不要有任何其他文字说明
- 没有的字段留空字符串
- 中文优先，尽量保留原文
- 设备名称要准确完整
- 型号、编号等不要有多余字符
- 从上下文中推断合理的信息"""

        user_prompt = f"""设备类别：{category_name}

OCR识别文本：
{ocr_text}

请提取设备信息，返回JSON格式。"""

        start_time = time.time()
        
        url = llm_config['base_url'].rstrip('/') + '/chat/completions'
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {llm_config["api_key"]}'
        }
        payload = {
            'model': llm_config['model'],
            'messages': [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            'temperature': 0.3,
            'max_tokens': 1000
        }
        
        response = requests.post(url, headers=headers, json=payload, timeout=60)
        response.raise_for_status()
        data = response.json()
        
        elapsed = time.time() - start_time
        content = data['choices'][0]['message']['content'].strip()
        
        content = re.sub(r'^```json\s*', '', content)
        content = re.sub(r'\s*```$', '', content)
        
        try:
            result = json.loads(content)
        except json.JSONDecodeError:
            json_match = re.search(r'\{[\s\S]*\}', content)
            if json_match:
                result = json.loads(json_match.group())
            else:
                return None, elapsed
        
        standardized = {
            'device_name': str(result.get('device_name', '')).strip(),
            'model': str(result.get('model', '')).strip(),
            'factory_number': str(result.get('factory_number', '')).strip(),
            'device_number': str(result.get('device_number', '')).strip(),
            'manufacturer': str(result.get('manufacturer', '')).strip(),
            'cert_date': str(result.get('cert_date', '')).strip(),
            'remark': str(result.get('remark', '')).strip(),
            'responsible_person': str(result.get('responsible_person', '')).strip(),
            'confidence': {'llm_based': 0.85}
        }
        
        return standardized, elapsed
        
    except Exception as e:
        print(f'LLM call failed: {e}')
        import traceback
        traceback.print_exc()
        return None, 0

COLORS = {
    'inspected': 'FF92D050',
    'brought_back': 'FF00B0F0',
    'returned': 'FFFF0000'
}

COLUMN_WIDTHS = {
    'A': 5.125, 'B': 14.25, 'C': 33.375, 'D': 18.75, 'E': 32.75,
    'F': 14.125, 'G': 35.375, 'H': 49.25, 'I': 10.0
}

ROW_HEIGHTS = {
    1: 27.0, 2: 25.0, 3: 25.0, 4: 25.0, 5: 25.0
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

COMMON_DEVICE_NAMES = [
    '压力表', '压力变送器', '压力传感器', '数字压力表', '精密压力表',
    '温度计', '温湿度计', '温湿度记录仪', '数字温度计', '铂电阻',
    '天平', '电子天平', '分析天平', '电子秤', '台秤', '地磅',
    '卡尺', '游标卡尺', '数显卡尺', '千分尺', '螺旋测微器',
    '万用表', '数字万用表', '钳形表', '兆欧表', '绝缘电阻表',
    '示波器', '数字示波器', '模拟示波器',
    '电源', '直流电源', '直流稳压电源', '交流电源',
    '电阻箱', '直流电阻箱', '交流电阻箱',
    '信号发生器', '函数信号发生器', '任意波形发生器',
    'PH计', '酸度计', '电导率仪', '电导仪',
    '分光光度计', '紫外可见分光光度计', '原子吸收分光光度计',
    '色谱仪', '气相色谱仪', '液相色谱仪',
    '流量计', '电磁流量计', '涡街流量计', '涡轮流量计',
    '液位计', '磁翻板液位计', '雷达液位计', '超声波液位计',
    '扭矩扳手', '扭力扳手', '数显扭矩扳手',
    '硬度计', '洛氏硬度计', '布氏硬度计', '维氏硬度计',
    '粗糙度仪', '表面粗糙度仪',
    '测厚仪', '超声波测厚仪', '涂层测厚仪',
    '探伤仪', '超声波探伤仪', '磁粉探伤仪',
    '试验机', '万能试验机', '拉力试验机', '压力试验机',
    '转速表', '声级计', '照度计', '风速仪', '尘埃粒子计数器'
]

COMMON_MANUFACTURERS = [
    '上海自动化仪表', '上海仪表', '北京康斯特', '北京时代', '深圳华仪',
    '福禄克', 'FLUKE',
    '优利德', 'UNI-T',
    '胜利', 'VICTOR',
    '天正', '天正电气', '天正仪表',
    '雷磁', '上海雷磁', '仪电科学',
    '梅特勒', '梅特勒托利多', 'METTLER TOLEDO',
    '赛多利斯', 'SARTORIUS',
    '岛津', 'SHIMADZU',
    '安捷伦', 'Agilent', '是德科技', 'Keysight',
    '横河', '横河电机', 'YOKOGAWA',
    '哈希', 'HACH',
    '魏德米勒', 'Weidmuller',
    '施耐德', 'Schneider',
    '西门子', 'SIEMENS',
    '欧姆龙', 'OMRON',
    '三菱', 'MITSUBISHI',
    '日立', 'HITACHI',
    '松下', 'Panasonic',
    '富士', 'FUJI',
    '得力', 'deli',
    '长城', '长城精工',
    '上工', '哈量', '成量', '青量'
]

OCR_CORRECTIONS = {
    'O': '0', 'o': '0', 'Q': '0', 'D': '0',
    'l': '1', 'I': '1', '|': '1', 'i': '1',
    'Z': '2', 'z': '2',
    'S': '5', 's': '5',
    'G': '6', 'b': '6',
    'T': '7',
    'B': '8',
    'g': '9', 'q': '9',
}

PUNCTUATION_CORRECTIONS = {
    '，': ':', '；': ':', '：': ':',
    '（': '(', '）': ')',
    '－': '-', '—': '-',
    '·': '.', '．': '.',
    '／': '/',
    'rn': 'm', 'cl': 'd',
}

def correct_ocr_text(text):
    if not text:
        return text
    for wrong, right in PUNCTUATION_CORRECTIONS.items():
        text = text.replace(wrong, right)
    return text

def correct_alnum(text):
    if not text:
        return text
    result = []
    for ch in text:
        if ch in OCR_CORRECTIONS:
            result.append(OCR_CORRECTIONS[ch])
        else:
            result.append(ch)
    return ''.join(result)

def calculate_text_quality(text):
    if not text:
        return 0
    score = 0
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if not lines:
        return 0
    score += min(len(lines) * 5, 50)
    total_len = sum(len(l) for l in lines)
    avg_len = total_len / len(lines) if lines else 0
    score += min(avg_len * 2, 20)
    chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', text))
    score += min(chinese_chars * 0.5, 20)
    has_keywords = any(k in text for k in ['名称', '型号', '编号', '厂家', 'SN', '出厂', '规格'])
    if has_keywords:
        score += 10
    digits = len(re.findall(r'\d', text))
    score += min(digits * 0.3, 10)
    return score

def preprocess_enhanced(img):
    if len(img.shape) == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        gray = img.copy()
    
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    return enhanced

def preprocess_denoise(img):
    if len(img.shape) == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        gray = img.copy()
    
    denoised = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
    return denoised

def preprocess_threshold(img):
    if len(img.shape) == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        gray = img.copy()
    
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return thresh

def preprocess_sharpen(img):
    if len(img.shape) == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        gray = img.copy()
    
    kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
    sharpened = cv2.filter2D(gray, -1, kernel)
    return sharpened

def run_ocr_easyocr(img):
    if ocr_reader is None:
        return [], 0
    
    try:
        results = ocr_reader.readtext(img)
        lines = []
        for (bbox, text, confidence) in results:
            lines.append((text, confidence))
        return lines, len(results)
    except Exception as e:
        print(f'EasyOCR error: {e}')
        return [], 0

def ocr_image(image_bytes):
    start = time.time()
    try:
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None:
            return '', 0, 0, 'image_decode_failed'
    except Exception as e:
        print(f'Image decode error: {e}')
        return '', 0, 0, 'image_decode_failed'
    
    decode_time = time.time() - start
    
    height, width = img.shape[:2]
    if max(height, width) > 3000:
        scale = 3000 / max(height, width)
        img = cv2.resize(img, (int(width * scale), int(height * scale)), interpolation=cv2.INTER_AREA)
    
    preprocess_start = time.time()
    strategies = [
        ('original', lambda x: cv2.cvtColor(x, cv2.COLOR_BGR2GRAY) if len(x.shape)==3 else x),
        ('enhanced', preprocess_enhanced),
        ('denoise', preprocess_denoise),
        ('threshold', preprocess_threshold),
        ('sharpen', preprocess_sharpen),
    ]
    
    best_text = ''
    best_score = 0
    best_lines = []
    
    for name, func in strategies:
        try:
            processed = func(img)
            lines, count = run_ocr_easyocr(processed)
            text = '\n'.join([l[0] for l in lines]) if lines else ''
            score = calculate_text_quality(text)
            
            if score > best_score:
                best_score = score
                best_text = text
                best_lines = lines
                
            if score >= 80:
                break
        except Exception as e:
            print(f'Strategy {name} failed: {e}')
    
    if best_score < 30:
        try:
            h, w = img.shape[:2]
            if max(h, w) < 1500:
                scale = 1500 / max(h, w)
                resized = cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_CUBIC)
                processed = preprocess_enhanced(resized)
                lines, count = run_ocr_easyocr(processed)
                text = '\n'.join([l[0] for l in lines]) if lines else ''
                score = calculate_text_quality(text)
                if score > best_score:
                    best_score = score
                    best_text = text
                    best_lines = lines
        except Exception as e:
            print(f'Resize strategy failed: {e}')
    
    ocr_time = time.time() - preprocess_start
    
    corrected_text = correct_ocr_text(best_text)
    
    if not corrected_text.strip():
        return corrected_text, ocr_time, decode_time, 'no_text_detected'
    
    return corrected_text, ocr_time, decode_time, None

def extract_device_info(text):
    info = {
        'device_name': '',
        'model': '',
        'factory_number': '',
        'device_number': '',
        'manufacturer': '',
        'confidence': {}
    }
    
    if not text:
        return info
    
    text = correct_ocr_text(text)
    
    normalized_lines = []
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue
        line = re.sub(r'\s+', ' ', line)
        normalized_lines.append(line)
    normalized_text = '\n'.join(normalized_lines)
    
    label_patterns = [
        (r'设备名称[：:\s]+([^\n\r]+)', 'device_name', 0.9),
        (r'产品名称[：:\s]+([^\n\r]+)', 'device_name', 0.85),
        (r'仪器名称[：:\s]+([^\n\r]+)', 'device_name', 0.85),
        (r'名\s*称[：:\s]+([^\n\r]+)', 'device_name', 0.7),
        (r'规格型号[：:\s]+([^\n\r]+)', 'model', 0.9),
        (r'型\s*号[：:\s]+([^\n\r]+)', 'model', 0.8),
        (r'型\s*式[：:\s]+([^\n\r]+)', 'model', 0.6),
        (r'Model[：:\s]+([^\n\r]+)', 'model', 0.75),
        (r'Type[：:\s]+([^\n\r]+)', 'model', 0.6),
        (r'出厂编号[：:\s]+([^\n\r]+)', 'factory_number', 0.95),
        (r'出厂号[：:\s]+([^\n\r]+)', 'factory_number', 0.85),
        (r'序\s*号[：:\s]+([^\n\r]+)', 'factory_number', 0.7),
        (r'编\s*号[：:\s]+([^\n\r]+)', 'factory_number', 0.6),
        (r'Serial[：:\s]+([^\n\r]+)', 'factory_number', 0.8),
        (r'SN[：:\s]+([A-Za-z0-9\-]+)', 'factory_number', 0.85),
        (r'sn[：:\s]+([A-Za-z0-9\-]+)', 'factory_number', 0.8),
        (r'No[.．][：:\s]+([A-Za-z0-9\-]+)', 'factory_number', 0.7),
        (r'设备编号[：:\s]+([^\n\r]+)', 'device_number', 0.9),
        (r'设备号[：:\s]+([^\n\r]+)', 'device_number', 0.8),
        (r'资产编号[：:\s]+([^\n\r]+)', 'device_number', 0.75),
        (r'生产厂家[：:\s]+([^\n\r]+)', 'manufacturer', 0.95),
        (r'制造厂家[：:\s]+([^\n\r]+)', 'manufacturer', 0.9),
        (r'厂\s*家[：:\s]+([^\n\r]+)', 'manufacturer', 0.8),
        (r'制造商[：:\s]+([^\n\r]+)', 'manufacturer', 0.85),
        (r'品\s*牌[：:\s]+([^\n\r]+)', 'manufacturer', 0.7),
        (r'Manufacturer[：:\s]+([^\n\r]+)', 'manufacturer', 0.8),
        (r'Made by[：:\s]+([^\n\r]+)', 'manufacturer', 0.7),
    ]
    
    for pattern, key, confidence in label_patterns:
        match = re.search(pattern, normalized_text, re.IGNORECASE)
        if match:
            value = match.group(1).strip()
            value = re.sub(r'[，。；、\s]+$', '', value)
            value = value.strip()
            if value and len(value) > 0:
                if not info[key] or confidence > info['confidence'].get(key, 0):
                    info[key] = value
                    info['confidence'][key] = confidence
    
    if not info['device_name']:
        for name in COMMON_DEVICE_NAMES:
            if name in normalized_text:
                info['device_name'] = name
                info['confidence']['device_name'] = 0.6
                break
    
    if not info['device_name']:
        patterns = [
            r'([\u4e00-\u9fff]{2,6}(表|计|仪|器|机|秤|天平|卡尺|千分尺|扳手|探头|传感器|变送器))',
        ]
        for pattern in patterns:
            matches = re.findall(pattern, normalized_text)
            if matches:
                info['device_name'] = matches[0][0] + matches[0][1]
                info['confidence']['device_name'] = 0.4
                break
    
    if not info['manufacturer']:
        for manu in COMMON_MANUFACTURERS:
            if manu in normalized_text:
                info['manufacturer'] = manu
                info['confidence']['manufacturer'] = 0.7
                break
    
    if not info['factory_number']:
        sn_patterns = [
            r'[A-Z]{2,4}[-_]?\d{5,12}',
            r'\d{8,15}',
            r'[A-Za-z]{1,3}\d{6,12}[A-Za-z0-9]{0,4}',
        ]
        for pattern in sn_patterns:
            matches = re.findall(pattern, normalized_text)
            for match in matches:
                skip = False
                for key in ['device_name', 'model', 'manufacturer']:
                    if info[key] and match in info[key]:
                        skip = True
                        break
                if not skip and len(match) >= 6:
                    info['factory_number'] = match
                    info['confidence']['factory_number'] = 0.5
                    break
            if info['factory_number']:
                break
    
    if not info['model']:
        model_patterns = [
            r'[A-Za-z][A-Za-z0-9]{0,3}[-_]?\d{2,5}[A-Za-z0-9]{0,6}',
            r'[\u4e00-\u9fffA-Za-z]\d{2,5}[-]?[A-Za-z0-9]{0,4}',
        ]
        for pattern in model_patterns:
            matches = re.findall(pattern, normalized_text)
            for match in matches:
                if len(match) >= 3 and not info['model']:
                    skip = False
                    for key in ['device_name', 'factory_number', 'manufacturer']:
                        if info[key] and match in info[key]:
                            skip = True
                            break
                    if not skip:
                        info['model'] = match
                        info['confidence']['model'] = 0.4
                        break
            if info['model']:
                break
    
    for key in ['device_name', 'model', 'factory_number', 'device_number', 'manufacturer']:
        if info[key]:
            info[key] = info[key].strip()
            info[key] = re.sub(r'\s+', ' ', info[key])
    
    for key in ['model', 'factory_number', 'device_number']:
        if info[key]:
            corrected = correct_alnum(info[key])
            if corrected != info[key]:
                info[key] = corrected
    
    return info

def merge_device_info(regex_info, llm_info):
    if llm_info is None:
        return regex_info
    
    merged = dict(regex_info)
    
    all_fields = ['device_name', 'model', 'factory_number', 'device_number', 
                  'manufacturer', 'cert_date', 'remark', 'responsible_person']
    
    for field in all_fields:
        llm_value = llm_info.get(field, '').strip()
        regex_value = regex_info.get(field, '').strip() if isinstance(regex_info, dict) else ''
        
        if llm_value and not regex_value:
            merged[field] = llm_value
        elif llm_value and regex_value:
            if field in ['remark', 'cert_date', 'responsible_person']:
                merged[field] = llm_value
            elif len(llm_value) >= len(regex_value):
                merged[field] = llm_value
    
    if 'confidence' not in merged:
        merged['confidence'] = {}
    merged['confidence']['llm_used'] = True
    
    return merged

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/config', methods=['GET'])
def get_config():
    return jsonify({
        'llm_enabled': llm_config['enabled'],
        'llm_model': llm_config['model'],
        'ocr_engine': ocr_engine,
        'max_file_size_mb': MAX_FILE_SIZE // 1024 // 1024,
        'allowed_extensions': list(ALLOWED_EXTENSIONS)
    })

@app.route('/upload', methods=['POST'])
def upload_image():
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件', 'error_code': 'no_file'}), 400
    
    file = request.files['file']
    category = request.form.get('category', 'inspected')
    
    if not file or file.filename == '':
        return jsonify({'error': '文件名为空', 'error_code': 'empty_filename'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': '不支持的文件格式，请上传图片文件', 'error_code': 'invalid_file_type'}), 400
    
    try:
        image_bytes = file.read()
        if len(image_bytes) == 0:
            return jsonify({'error': '文件内容为空', 'error_code': 'empty_file'}), 400
        if len(image_bytes) > MAX_FILE_SIZE:
            return jsonify({'error': f'文件大小超过限制（最大{MAX_FILE_SIZE//1024//1024}MB）', 'error_code': 'file_too_large'}), 400
        
        filename = str(uuid.uuid4()) + '_' + file.filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        with open(filepath, 'wb') as f:
            f.write(image_bytes)
        
        init_ocr()
        
        if ocr_reader is None:
            return jsonify({'error': 'OCR引擎初始化失败，请检查环境', 'error_code': 'ocr_init_failed'}), 500
        
        text, ocr_time, decode_time, error_code = ocr_image(image_bytes)
        
        extract_start = time.time()
        regex_info = extract_device_info(text)
        extract_time = time.time() - extract_start
        
        llm_info = None
        llm_time = 0
        if text and text.strip():
            llm_start = time.time()
            llm_info, llm_time = call_llm_extract(text, category)
            llm_time = time.time() - llm_start
        
        info = merge_device_info(regex_info, llm_info)
        
        total_time = ocr_time + decode_time + extract_time + llm_time
        
        info['category'] = category
        info['ocr_text'] = text
        info['filename'] = filename
        info['ocr_engine'] = ocr_engine
        info['llm_used'] = llm_info is not None
        info['llm_model'] = llm_config['model'] if llm_info is not None else ''
        info['success'] = error_code is None
        if error_code:
            info['error_code'] = error_code
        info['timing'] = {
            'decode_time': round(decode_time, 3),
            'ocr_time': round(ocr_time, 3),
            'extract_time': round(extract_time, 3),
            'llm_time': round(llm_time, 3),
            'total_time': round(total_time, 3)
        }
        
        return jsonify(info)
    
    except Exception as e:
        print(f'Upload error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'服务器处理出错：{str(e)}', 'error_code': 'server_error'}), 500

@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    start_time = time.time()
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': '请求数据为空', 'error_code': 'empty_request'}), 400
        
        company_info = data.get('company_info', {})
        device_list = data.get('devices', [])
        
        wb = Workbook()
        ws = wb.active
        ws.title = '空白表'
        
        for col, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width
        
        for row, height in ROW_HEIGHTS.items():
            ws.row_dimensions[row].height = height
        
        header_font = Font(name='华文仿宋', size=11, bold=False)
        title_font = Font(name='华文仿宋', size=14, bold=False)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        color_fills = {
            'inspected': PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid'),
            'brought_back': PatternFill(start_color='FF00B0F0', end_color='FF00B0F0', fill_type='solid'),
            'returned': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        }
        
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws.merge_cells('A4:F4')
        
        ws['A1'] = company_info.get('company', '公司-下场单')
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        ws['G1'] = f"业  务  员：{company_info.get('salesperson', '')}"
        ws['G1'].font = header_font
        ws['G1'].alignment = left_align
        
        ws['H1'] = f"下场负责人：{company_info.get('field_leader', '')}"
        ws['H1'].font = header_font
        ws['H1'].alignment = left_align
        
        ws['I1'] = '颜色示例'
        ws['I1'].font = header_font
        ws['I1'].alignment = center_align
        
        ws['A2'] = f"单位名称：{company_info.get('company_name', '')}"
        ws['A2'].font = header_font
        ws['A2'].alignment = left_align
        
        ws['G2'] = f"联  系  人：{company_info.get('contact_person', '')}"
        ws['G2'].font = header_font
        ws['G2'].alignment = left_align
        
        ws['H2'] = f"联系方式：{company_info.get('contact_phone', '')}"
        ws['H2'].font = header_font
        ws['H2'].alignment = left_align
        
        ws['I2'] = '带回'
        ws['I2'].font = header_font
        ws['I2'].alignment = center_align
        ws['I2'].fill = color_fills['brought_back']
        
        ws['A3'] = f"单位地址：{company_info.get('address', '')}"
        ws['A3'].font = header_font
        ws['A3'].alignment = left_align
        
        ws['G3'] = f"下场日期：{company_info.get('field_date', '')}"
        ws['G3'].font = header_font
        ws['G3'].alignment = left_align
        
        ws['H3'] = f"下场人员：{company_info.get('field_personnel', '')}"
        ws['H3'].font = header_font
        ws['H3'].alignment = left_align
        
        ws['I3'] = '退检'
        ws['I3'].font = header_font
        ws['I3'].alignment = center_align
        ws['I3'].fill = color_fills['returned']
        
        ws['A4'] = f"委托单备注：{company_info.get('remark', '/')}"
        ws['A4'].font = header_font
        ws['A4'].alignment = left_align
        
        device_count = len(device_list)
        brought_back_count = sum(1 for d in device_list if d.get('category') == 'brought_back')
        
        ws['G4'] = f"设备总数：{device_count}"
        ws['G4'].font = header_font
        ws['G4'].alignment = left_align
        
        ws['H4'] = f"带回数量：{brought_back_count}"
        ws['H4'].font = header_font
        ws['H4'].alignment = left_align
        
        ws['I4'] = '已检'
        ws['I4'].font = header_font
        ws['I4'].alignment = center_align
        ws['I4'].fill = color_fills['inspected']
        
        headers = ['序号', '证书日期', '设备名称', '规格型号', '出厂编号', '设备编号', '生产厂家', '设备备注（特殊要求、校准信息、地点地点、温湿度等）', '负责人']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, device in enumerate(device_list, 6):
            category = device.get('category', 'inspected')
            fill = color_fills.get(category, PatternFill(fill_type=None))
            
            cells = [
                (1, row_idx - 5, center_align),
                (2, device.get('cert_date', ''), center_align),
                (3, device.get('device_name', ''), left_align),
                (4, device.get('model', ''), center_align),
                (5, device.get('factory_number', ''), center_align),
                (6, device.get('device_number', ''), center_align),
                (7, device.get('manufacturer', ''), center_align),
                (8, device.get('remark', ''), left_align),
                (9, device.get('responsible_person', ''), center_align)
            ]
            
            for col, value, align in cells:
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = header_font
                cell.alignment = align
                if col == 9:
                    cell.fill = PatternFill(fill_type=None)
                else:
                    cell.fill = fill
                cell.border = thin_border
        
        date_part = company_info.get('field_date', '未命名')
        name_part = company_info.get('company_name', '未命名')
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', name_part)
        output_filename = f"{date_part}_{safe_name}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        wb.save(output_path)
        
        generate_time = round(time.time() - start_time, 3)
        
        return jsonify({'filename': output_filename, 'generate_time': generate_time, 'success': True})
    
    except Exception as e:
        print(f'Excel generate error: {e}')
        return jsonify({'error': f'生成Excel失败：{str(e)}', 'error_code': 'excel_gen_error'}), 500

@app.route('/download/<filename>')
def download_file(filename):
    try:
        return send_from_directory(app.config['OUTPUT_FOLDER'], filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': '文件不存在或下载失败', 'error_code': 'download_error'}), 404

@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({'error': f'文件大小超过限制（最大{MAX_FILE_SIZE//1024//1024}MB）', 'error_code': 'file_too_large'}), 413

def create_directories():
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

create_directories()

if __name__ == '__main__':
    create_directories()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=DEBUG_MODE, host='0.0.0.0', port=port)